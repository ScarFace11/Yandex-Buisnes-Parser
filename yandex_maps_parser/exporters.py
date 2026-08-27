"""
File output: CSV, JSON, Excel (.xlsx), and HTML map.
"""
import csv
import json
import os
import re
import threading
from collections import Counter
from datetime import datetime

import openpyxl
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from .constants import (
    CSV_FIELDS,
    HEADER_LABELS,
    COL_WIDTHS,
    SOCIAL_COLORS,
    SOCIAL_BADGE_COLORS,
    SOCIAL_LABELS,
    KNOWN_PLATFORMS,
    QUERY_COLORS,
)
from . import state


# ── Path resolution ───────────────────────────────────────────

def _resolve(base: str) -> dict[str, str]:
    os.makedirs(state.OUTPUT_DIR, exist_ok=True)
    return {
        "csv":  os.path.join(state.OUTPUT_DIR, f"{base}.csv"),
        "json": os.path.join(state.OUTPUT_DIR, f"{base}.json"),
        "jsonl": os.path.join(state.OUTPUT_DIR, f"{base}.jsonl"),
        "xlsx": os.path.join(state.OUTPUT_DIR, f"{base}.xlsx"),
        "map":  os.path.join(state.OUTPUT_DIR, f"{base}_map.html"),
    }


def load_jsonl(path: str) -> list[dict]:
    """Read records from the crash-safe JSONL sidecar (one JSON object per line)."""
    records: list[dict] = []
    if not os.path.exists(path):
        return records
    try:
        with open(path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    records.append(json.loads(line))
    except Exception:
        pass
    return records


def record_key(rec: dict) -> str:
    """
    Stable dedup key for a business record.
    Priority: yandex_maps_url -> phone digits -> normalized name+address.
    """
    url = str(rec.get("yandex_maps_url") or "").strip()
    if url:
        return "u|" + url
    digits = re.sub(r"\D", "", str(rec.get("phone") or ""))
    if digits:
        if len(digits) == 11 and digits[0] == "8":
            digits = "7" + digits[1:]
        return "p|" + digits
    norm = lambda s: re.sub(r"\W+", "", (s or "").lower())
    return f"n|{norm(rec.get('name'))}|{norm(rec.get('address'))}"


def dedupe_records(records: list[dict]) -> list[dict]:
    """Deduplicate records using record_key()."""
    seen: set[str] = set()
    out: list[dict] = []
    for r in records:
        key = record_key(r)
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def load_existing_urls(csv_path: str) -> set[str]:
    ids: set[str] = set()
    if not os.path.exists(csv_path):
        return ids
    try:
        with open(csv_path, encoding="utf-8-sig") as f:
            for row in csv.DictReader(f):
                if u := row.get("yandex_maps_url", ""):
                    ids.add(u)
    except Exception:
        pass
    return ids


# ── CSV ───────────────────────────────────────────────────────

def save_csv(records: list[dict], path: str, append: bool) -> None:
    mode = "a" if append and os.path.exists(path) else "w"
    with open(path, mode, newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=CSV_FIELDS, extrasaction="ignore")
        if mode == "w":
            w.writeheader()
        w.writerows(records)


# ── JSON ──────────────────────────────────────────────────────

def save_json(records: list[dict], path: str, append: bool) -> None:
    existing: list[dict] = []
    if append and os.path.exists(path):
        try:
            with open(path, encoding="utf-8") as f:
                existing = json.load(f)
        except Exception:
            pass
    with open(path, "w", encoding="utf-8") as f:
        json.dump(existing + records, f, ensure_ascii=False, indent=2)


# ── Excel shared styles ──────────────────────────────────────

_THIN        = Side(style="thin",   color="CCCCCC")
_THICK       = Side(style="medium", color="AAAAAA")
CELL_BORDER  = Border(left=_THIN, right=_THIN, top=_THIN,  bottom=_THIN)
HEAD_BORDER  = Border(left=_THIN, right=_THIN, top=_THICK, bottom=_THICK)
HDR_FILL     = PatternFill("solid", fgColor="1A6B3C")
HDR_FONT     = Font(bold=True, color="FFFFFF", size=10)
REV_FILL     = PatternFill("solid", fgColor="E8F5E9")
ALT_FILL     = PatternFill("solid", fgColor="F8F8F8")
LINK_FONT    = Font(color="1155CC", underline="single", size=10)
URL_RE       = re.compile(r"https?://", re.I)

_EXCEL_LOCK = threading.Lock()  # guards _Excel state below


def _write_row(ws, ri: int, record: dict) -> None:
    """Write a single business record to worksheet row ri (1=header, 2+=data)."""
    alt = ri % 2 == 0
    for ci, field in enumerate(CSV_FIELDS, 1):
        val  = record.get(field, "")
        cell = ws.cell(row=ri, column=ci)

        if field == "reviewed":
            cell.fill  = REV_FILL
            cell.value = val
        elif field in SOCIAL_COLORS and isinstance(val, str) and URL_RE.match(val):
            cell.fill      = PatternFill("solid", fgColor=SOCIAL_COLORS[field])
            cell.hyperlink = val
            cell.value     = val
            cell.font      = Font(color="1155CC", underline="single", size=10)
        elif field in SOCIAL_COLORS and not val:
            cell.value = ""
        elif field in ("aggregator_url", "yandex_maps_url") and isinstance(val, str) and URL_RE.match(val):
            cell.hyperlink = val
            cell.value     = val
            cell.font      = Font(color="1155CC", underline="single", size=10)
            if alt:
                cell.fill = ALT_FILL
        elif field == "reviews":
            try:
                cell.value = int(str(val).replace("\xa0", "").replace(" ", "")) if val not in ("", None) else 0
            except (TypeError, ValueError):
                cell.value = 0
            cell.number_format = '#,##0'
            if alt:
                cell.fill = ALT_FILL
        else:
            cell.value = val
            if alt and field not in SOCIAL_COLORS:
                cell.fill = ALT_FILL

        cell.alignment = Alignment(vertical="top", wrap_text=(field == "description"))
        cell.border    = CELL_BORDER


def _write_headers(ws) -> None:
    """Write styled header row to an empty worksheet."""
    for ci, field in enumerate(CSV_FIELDS, 1):
        c = ws.cell(row=1, column=ci, value=HEADER_LABELS.get(field, field))
        c.font      = HDR_FONT
        c.fill      = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = HEAD_BORDER
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "B2"
    for ci, field in enumerate(CSV_FIELDS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(field, 16)
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"
    ws.print_title_rows = "1:1"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(CSV_FIELDS))}1"


# ── Excel incremental (web mode) ─────────────────────────────
#
# The workbook stays open in memory during the run.  Each enriched record
# is appended immediately so the user can download a partially-complete
# file at any time.  Thread-safe via _EXCEL_LOCK.

class _ExcelState:
    """Mutable container for the in-progress workbook."""
    __slots__ = ("wb", "path", "open")

    def __init__(self):
        self.wb: openpyxl.Workbook | None = None
        self.path: str = ""
        self.open: bool = False


_excel = _ExcelState()


def _init_excel(path: str) -> None:
    """Create a fresh workbook with headers and prepare for incremental writes."""
    with _EXCEL_LOCK:
        if _excel.open:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "\u0411\u0438\u0437\u043d\u0435\u0441\u044b"
        ws.sheet_view.showGridLines = False
        _write_headers(ws)
        wb.save(path)
        _excel.wb   = wb
        _excel.path = path
        _excel.open = True


def _append_excel(record: dict) -> None:
    """Append one record to the in-progress workbook (thread-safe)."""
    with _EXCEL_LOCK:
        if not _excel.open or _excel.wb is None:
            return
        ws = _excel.wb.active
        ri = ws.max_row + 1
        _write_row(ws, ri, record)
        # Update autofilter range
        ws.auto_filter.ref = f"A1:{get_column_letter(len(CSV_FIELDS))}{ws.max_row}"
        # Save every record so the file is always downloadable
        try:
            _excel.wb.save(_excel.path)
        except Exception:
            pass


def _finalize_excel(records: list[dict]) -> None:
    """Add table, conditional formatting, legend/help/stats, then close."""
    with _EXCEL_LOCK:
        if not _excel.open or _excel.wb is None:
            return
        wb = _excel.wb
        ws = wb.active

        # Table + autofilter
        table_ref = f"A1:{get_column_letter(len(CSV_FIELDS))}{max(ws.max_row, 2)}"
        if ws.max_row >= 2:
            # Clear existing tables (from incremental saves)
            for tbl_name in list(ws.tables.keys()):
                del ws.tables[tbl_name]
            table = Table(displayName="BusinessesTable", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium4",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=False,
                showColumnStripes=False,
            )
            ws.add_table(table)
        ws.auto_filter.ref = table_ref
        ws.freeze_panes = "B2"

        # Conditional formatting
        rating_col  = get_column_letter(CSV_FIELDS.index("rating")  + 1)
        reviews_col = get_column_letter(CSV_FIELDS.index("reviews") + 1)
        ws.conditional_formatting._cf_rules.clear()
        if ws.max_row >= 2:
            ws.conditional_formatting.add(
                f"{rating_col}2:{rating_col}{ws.max_row}",
                ColorScaleRule(
                    start_type="min", start_color="F8696B",
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                    end_type="max", end_color="63BE7B",
                ),
            )
            ws.conditional_formatting.add(
                f"{reviews_col}2:{reviews_col}{ws.max_row}",
                DataBarRule(start_type="min", end_type="max", color="5B9BD5", showValue=True),
            )

        # Extra sheets
        _fill_legend_sheet(wb.create_sheet("\u041b\u0435\u0433\u0435\u043d\u0434\u0430 \u0446\u0432\u0435\u0442\u043e\u0432"))
        _fill_help_sheet(wb.create_sheet("\u041a\u0430\u043a \u043f\u043e\u043b\u044c\u0437\u043e\u0432\u0430\u0442\u044c\u0441\u044f"))
        _fill_stats_sheet(wb.create_sheet("\u0421\u0442\u0430\u0442\u0438\u0441\u0442\u0438\u043a\u0430"), records, HDR_FILL, HDR_FONT, CELL_BORDER, ALT_FILL)

        wb.save(_excel.path)
        _excel.open = False
        _excel.wb   = None


# ── Excel (batch / CLI) ──────────────────────────────────────

def save_excel(all_records: list[dict], path: str) -> None:
    """Create a complete Excel file from scratch (used by CLI)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "\u0411\u0438\u0437\u043d\u0435\u0441\u044b"
    ws.sheet_view.showGridLines = False

    _write_headers(ws)

    for ri, record in enumerate(all_records, 2):
        _write_row(ws, ri, record)

    table_ref = f"A1:{get_column_letter(len(CSV_FIELDS))}{max(ws.max_row, 2)}"
    if ws.max_row >= 2:
        table = Table(displayName="BusinessesTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        ws.add_table(table)
    ws.auto_filter.ref = table_ref

    rating_col  = get_column_letter(CSV_FIELDS.index("rating")  + 1)
    reviews_col = get_column_letter(CSV_FIELDS.index("reviews") + 1)
    if ws.max_row >= 2:
        ws.conditional_formatting.add(
            f"{rating_col}2:{rating_col}{ws.max_row}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            ),
        )
        ws.conditional_formatting.add(
            f"{reviews_col}2:{reviews_col}{ws.max_row}",
            DataBarRule(start_type="min", end_type="max", color="5B9BD5", showValue=True),
        )

    _fill_legend_sheet(wb.create_sheet("\u041b\u0435\u0433\u0435\u043d\u0434\u0430 \u0446\u0432\u0435\u0442\u043e\u0432"))
    _fill_help_sheet(wb.create_sheet("\u041a\u0430\u043a \u043f\u043e\u043b\u044c\u0437\u043e\u0432\u0430\u0442\u044c\u0441\u044f"))
    _fill_stats_sheet(wb.create_sheet("\u0421\u0442\u0430\u0442\u0438\u0441\u0442\u0438\u043a\u0430"), all_records, HDR_FILL, HDR_FONT, CELL_BORDER, ALT_FILL)

    wb.save(path)


# ── Legend / Help / Stats sheets ──────────────────────────────

def _fill_legend_sheet(ws) -> None:
    brd  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 14

    for ci, label in enumerate(["\u041f\u043b\u0430\u0442\u0444\u043e\u0440\u043c\u0430", "\u0426\u0432\u0435\u0442 \u044f\u0447\u0435\u0439\u043a\u0438 \u043f\u0440\u0438 \u043d\u0430\u043b\u0438\u0447\u0438\u0438 \u0441\u0441\u044b\u043b\u043a\u0438", "Hex-\u043a\u043e\u0434"], 1):
        c = ws.cell(row=1, column=ci, value=label)
        c.font      = HDR_FONT
        c.fill      = HDR_FILL
        c.border    = brd
        c.alignment = Alignment(horizontal="center")

    labels = {
        "vk": "\u0412\u041a\u043e\u043d\u0442\u0430\u043a\u0442\u0435", "instagram": "Instagram", "facebook": "Facebook",
        "telegram": "Telegram", "youtube": "YouTube", "tiktok": "TikTok",
        "ok": "\u041e\u0434\u043d\u043e\u043a\u043b\u0430\u0441\u0441\u043d\u0438\u043a\u0438", "twitter": "Twitter / X", "whatsapp": "WhatsApp",
        "other_socials": "\u0414\u0440\u0443\u0433\u0438\u0435 \u0441\u043e\u0446\u0441\u0435\u0442\u0438",
    }
    for ri, (platform, hex_color) in enumerate(SOCIAL_COLORS.items(), 2):
        ws.cell(row=ri, column=1, value=labels.get(platform, platform)).border = brd
        c = ws.cell(row=ri, column=2, value="\u041d\u0430\u0439\u0434\u0435\u043d\u0430 \u0441\u0441\u044b\u043b\u043a\u0430")
        c.fill   = PatternFill("solid", fgColor=hex_color)
        c.border = brd
        ws.cell(row=ri, column=3, value=f"#{hex_color}").border = brd

    r = len(SOCIAL_COLORS) + 2
    ws.cell(r, 1, "(\u043f\u0443\u0441\u0442\u043e)").border          = brd
    ws.cell(r, 2, "\u0421\u0441\u044b\u043b\u043a\u0430 \u043d\u0435 \u043d\u0430\u0439\u0434\u0435\u043d\u0430").border = brd


def _fill_help_sheet(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 95

    title_fill = PatternFill("solid", fgColor="1A6B3C")
    title_font = Font(bold=True, color="FFFFFF", size=12)
    label_font = Font(bold=True, color="1A6B3C", size=10)

    ws["A1"] = "\u041a\u0430\u043a \u0440\u0430\u0431\u043e\u0442\u0430\u0442\u044c \u0441 \u0442\u0430\u0431\u043b\u0438\u0446\u0435\u0439"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["B1"].fill = title_fill
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 28

    tips = [
        ("\u041f\u043e\u0438\u0441\u043a", "\u041e\u0442\u043a\u0440\u043e\u0439\u0442\u0435 \u0444\u0438\u043b\u044c\u0442\u0440 \u0432 \u0437\u0430\u0433\u043e\u043b\u043e\u0432\u043a\u0435 \u043d\u0443\u0436\u043d\u043e\u0433\u043e \u0441\u0442\u043e\u043b\u0431\u0446\u0430 \u0438 \u0432\u0432\u0435\u0434\u0438\u0442\u0435 \u0442\u0435\u043a\u0441\u0442 \u0432 \u043f\u043e\u043b\u0435 \u00ab\u041f\u043e\u0438\u0441\u043a\u00bb. Excel \u043f\u043e\u043a\u0430\u0436\u0435\u0442 \u0442\u043e\u043b\u044c\u043a\u043e \u043f\u043e\u0434\u0445\u043e\u0434\u044f\u0449\u0438\u0435 \u0441\u0442\u0440\u043e\u043a\u0438."),
        ("\u0424\u0438\u043b\u044c\u0442\u0440 \u043f\u043e \u043e\u0442\u0437\u044b\u0432\u0430\u043c", "\u0412 \u0441\u0442\u043e\u043b\u0431\u0446\u0435 \u00ab\u041e\u0442\u0437\u044b\u0432\u043e\u0432\u00bb \u0432\u044b\u0431\u0435\u0440\u0438\u0442\u0435 \u00ab\u0427\u0438\u0441\u043b\u043e\u0432\u044b\u0435 \u0444\u0438\u043b\u044c\u0442\u0440\u044b\u00bb \u0438 \u0437\u0430\u0434\u0430\u0439\u0442\u0435 \u0443\u0441\u043b\u043e\u0432\u0438\u0435."),
        ("\u0424\u0438\u043b\u044c\u0442\u0440 \u043f\u043e \u0440\u0435\u0439\u0442\u0438\u043d\u0433\u0443", "\u0412 \u0441\u0442\u043e\u043b\u0431\u0446\u0435 \u00ab\u0420\u0435\u0439\u0442\u0438\u043d\u0433\u00bb \u043c\u043e\u0436\u043d\u043e \u0432\u044b\u0431\u0440\u0430\u0442\u044c \u0437\u043d\u0430\u0447\u0435\u043d\u0438\u044f \u0438\u043b\u0438 \u043f\u0440\u0438\u043c\u0435\u043d\u0438\u0442\u044c \u0447\u0438\u0441\u043b\u043e\u0432\u043e\u0435 \u0443\u0441\u043b\u043e\u0432\u0438\u0435."),
        ("\u0421\u043e\u0440\u0442\u0438\u0440\u043e\u0432\u043a\u0430", "\u041d\u0430\u0436\u043c\u0438\u0442\u0435 \u0441\u0442\u0440\u0435\u043b\u043a\u0443 \u0432 \u0437\u0430\u0433\u043e\u043b\u043e\u0432\u043a\u0435 \u0441\u0442\u043e\u043b\u0431\u0446\u0430 \u0438 \u0432\u044b\u0431\u0435\u0440\u0438\u0442\u0435 \u0441\u043e\u0440\u0442\u0438\u0440\u043e\u0432\u043a\u0443."),
        ("\u041d\u0435\u0441\u043a\u043e\u043b\u044c\u043a\u043e \u0443\u0441\u043b\u043e\u0432\u0438\u0439", "\u0424\u0438\u043b\u044c\u0442\u0440\u044b \u0440\u0430\u0437\u043d\u044b\u0445 \u0441\u0442\u043e\u043b\u0431\u0446\u043e\u0432 \u0440\u0430\u0431\u043e\u0442\u0430\u044e\u0442 \u043e\u0434\u043d\u043e\u0432\u0440\u0435\u043c\u0435\u043d\u043d\u043e."),
        ("\u0417\u0430\u043a\u0440\u0435\u043f\u043b\u0435\u043d\u0438\u0435", "\u0412\u0435\u0440\u0445\u043d\u044f\u044f \u0441\u0442\u0440\u043e\u043a\u0430 \u043e\u0441\u0442\u0430\u0451\u0442\u0441\u044f \u0432\u0438\u0434\u0438\u043c\u043e\u0439 \u043f\u0440\u0438 \u043f\u0440\u043e\u043a\u0440\u0443\u0442\u043a\u0435."),
    ]
    for row, (label, text) in enumerate(tips, 3):
        ws.cell(row, 1, label).font = label_font
        cell = ws.cell(row, 2, text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 34


def _fill_stats_sheet(ws, records, hfill, hfont, border, alt_fill) -> None:
    brd = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

    def hdr(row, col, val):
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        c           = ws.cell(row=row, column=col, value=val)
        c.font      = hfont
        c.fill      = hfill
        c.border    = brd
        c.alignment = Alignment(horizontal="center")

    def cell(row, col, val, alt=False):
        c = ws.cell(row=row, column=col, value=val)
        if alt:
            c.fill = alt_fill
        c.border = brd
        return c

    row = 1
    hdr(row, 1, "\U0001f4ca \u041e\u0431\u0449\u0430\u044f \u0441\u0442\u0430\u0442\u0438\u0441\u0442\u0438\u043a\u0430")
    row += 1
    for label, val in [
        ("\u0412\u0441\u0435\u0433\u043e \u043d\u0430\u0439\u0434\u0435\u043d\u043e",            len(records)),
        ("\u0421 \u043e\u043f\u0438\u0441\u0430\u043d\u0438\u0435\u043c",              sum(1 for r in records if r.get("description"))),
        ("\u0421 \u0440\u0435\u0439\u0442\u0438\u043d\u0433\u043e\u043c",              sum(1 for r in records if r.get("rating"))),
        ("\u0421 \u043a\u043e\u043b\u0438\u0447\u0435\u0441\u0442\u0432\u043e\u043c \u043e\u0442\u0437\u044b\u0432\u043e\u0432",    sum(1 for r in records if r.get("reviews") not in ("", None, 0))),
        ("\u0427\u0435\u0440\u0435\u0437 taplink/linktree",   sum(1 for r in records if r.get("aggregator_url"))),
        ("\u0421 \u043f\u0440\u043e\u0432\u0435\u0440\u043a\u043e\u0439 \u0441\u043e\u0446\u0441\u0435\u0442\u0435\u0439",     sum(1 for r in records if r.get("socials_valid"))),
    ]:
        cell(row, 1, label, row % 2 == 0)
        cell(row, 2, val,   row % 2 == 0)
        row += 1

    row += 1
    hdr(row, 1, "\U0001f50d \u041f\u043e \u0437\u0430\u043f\u0440\u043e\u0441\u0430\u043c")
    row += 1
    for q, cnt in Counter(r.get("query", "") for r in records).most_common():
        cell(row, 1, q,   row % 2 == 0)
        cell(row, 2, cnt, row % 2 == 0)
        row += 1

    row += 1
    hdr(row, 1, "\U0001f4f1 \u041f\u043e \u0441\u043e\u0446\u0441\u0435\u0442\u044f\u043c")
    row += 1
    labels = {
        "vk": "\u0412\u041a\u043e\u043d\u0442\u0430\u043a\u0442\u0435", "instagram": "Instagram", "facebook": "Facebook",
        "telegram": "Telegram", "youtube": "YouTube", "tiktok": "TikTok",
        "ok": "\u041e\u0434\u043d\u043e\u043a\u043b\u0430\u0441\u0441\u043d\u0438\u043a\u0438", "twitter": "Twitter/X", "whatsapp": "WhatsApp",
    }
    for p in KNOWN_PLATFORMS:
        cnt = sum(1 for r in records if r.get(p))
        if cnt:
            c1 = cell(row, 1, labels.get(p, p), row % 2 == 0)
            c1.fill = PatternFill("solid", fgColor=SOCIAL_COLORS.get(p, "FFFFFF"))
            cell(row, 2, cnt, row % 2 == 0)
            row += 1

    row += 1
    hdr(row, 1, "\U0001f3ea \u0422\u043e\u043f \u043a\u0430\u0442\u0435\u0433\u043e\u0440\u0438\u0439")
    row += 1
    cats: Counter = Counter()
    for r in records:
        for cat in r.get("category", "").split(","):
            if c2 := cat.strip():
                cats[c2] += 1
    for cat, cnt in cats.most_common(15):
        cell(row, 1, cat, row % 2 == 0)
        cell(row, 2, cnt, row % 2 == 0)
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12


# ── HTML map ──────────────────────────────────────────────────

def _popup_html(r: dict, idx: int) -> str:
    name     = r.get("name", "")
    category = r.get("category", "")
    address  = r.get("address", "")
    phone    = r.get("phone", "")
    rating   = r.get("rating", "")
    reviews  = r.get("reviews", "")
    maps_url = r.get("yandex_maps_url", "")
    agg_url  = r.get("aggregator_url", "")
    hours    = r.get("hours", "")

    badges = ""
    for p in KNOWN_PLATFORMS:
        url = r.get(p, "")
        if url:
            color = SOCIAL_BADGE_COLORS.get(p, "#888")
            label = SOCIAL_LABELS.get(p, p.upper())
            badges += (
                f'<a href="{url}" target="_blank" style="'
                f'display:inline-block;margin:2px 3px 2px 0;padding:2px 7px;'
                f'background:{color};color:#fff;border-radius:4px;'
                f'font-size:11px;font-weight:bold;text-decoration:none">'
                f'{label}</a>'
            )
    other = r.get("other_socials", "")
    if other:
        for u in other.split(", "):
            u = u.strip()
            if u:
                badges += (
                    f'<a href="{u}" target="_blank" style="'
                    f'display:inline-block;margin:2px 3px 2px 0;padding:2px 7px;'
                    f'background:#9C27B0;color:#fff;border-radius:4px;'
                    f'font-size:11px;font-weight:bold;text-decoration:none">\u2026</a>'
                )

    rating_str = ""
    if rating:
        stars = "\u2605" * round(float(rating)) + "\u2606" * (5 - round(float(rating)))
        rating_str = (
            f'<div style="margin:4px 0;color:#f5a623;font-size:13px">'
            f'{stars} <span style="color:#555;font-size:12px">'
            f'{rating}{(" \u00b7 " + reviews + " \u043e\u0442\u0437.") if reviews else ""}</span></div>'
        )

    rows = ""
    if category:  rows += f'<div style="color:#888;font-size:11px;margin-bottom:3px">{category}</div>'
    if rating_str: rows += rating_str
    if address:   rows += f'<div style="margin:3px 0;font-size:12px">\U0001f4cd {address}</div>'
    if phone:     rows += f'<div style="margin:3px 0;font-size:12px">\U0001f4de <a href="tel:{phone}">{phone}</a></div>'
    if hours:     rows += f'<div style="margin:3px 0;font-size:12px;color:#555">\U0001f550 {hours}</div>'
    if agg_url:
        rows += (
            f'<div style="margin:4px 0;font-size:12px">'
            f'\U0001f517 <a href="{agg_url}" target="_blank" style="color:#FF6B35">Taplink / Linktree</a></div>'
        )
    if badges:    rows += f'<div style="margin:6px 0">{badges}</div>'
    if maps_url:
        rows += (
            f'<div style="margin:6px 0">'
            f'<a href="{maps_url}" target="_blank" style="'
            f'display:inline-block;padding:4px 10px;background:#FF0000;'
            f'color:#fff;border-radius:4px;font-size:12px;text-decoration:none">'
            f'\u041e\u0442\u043a\u0440\u044b\u0442\u044c \u043d\u0430 \u042f\u043d\u0434\u0435\u043a\u0441.\u041a\u0430\u0440\u0442\u0430\u0445 \u2197</a></div>'
        )

    return (
        f'<div style="font-family:Arial,sans-serif;min-width:220px;max-width:280px">'
        f'<b style="font-size:14px">{name}</b>'
        f'{rows}</div>'
    )


def _folium_color_to_hex(name: str) -> str:
    mapping = {
        "blue":      "#4285F4", "red":       "#EA4335", "green":     "#34A853",
        "purple":    "#9C27B0", "orange":    "#FF9800", "darkblue":  "#1A237E",
        "darkred":   "#B71C1C", "darkgreen": "#1B5E20", "cadetblue": "#006064",
        "lightred":  "#EF9A9A",
    }
    return mapping.get(name, "#4285F4")


def save_map(records: list[dict], path: str, center_lat: float, center_lon: float) -> None:
    import folium
    from folium.plugins import HeatMap, MarkerCluster

    m = folium.Map(
        location=[center_lat, center_lon],
        zoom_start=12,
        tiles="OpenStreetMap",
        prefer_canvas=True,
    )

    queries  = list(dict.fromkeys(r.get("query", "") for r in records))
    q_color  = {q: QUERY_COLORS[i % len(QUERY_COLORS)] for i, q in enumerate(queries)}
    q_clusters = {}
    for q in queries:
        fg = folium.FeatureGroup(name=f"\U0001f50d {q}", show=True)
        q_clusters[q] = (fg, MarkerCluster().add_to(fg))
        fg.add_to(m)

    heat_fg = folium.FeatureGroup(name="\U0001f321 \u0422\u0435\u043f\u043b\u043e\u0432\u0430\u044f \u043a\u0430\u0440\u0442\u0430", show=False)
    heat_fg.add_to(m)

    heat_points = []
    for idx, r in enumerate(records):
        lat = r.get("lat")
        lon = r.get("lon")
        try:
            lat = float(lat)
            lon = float(lon)
        except (TypeError, ValueError):
            continue

        heat_points.append([lat, lon])
        q     = r.get("query", "")
        color = q_color.get(q, "blue")

        icon_color = "orange" if r.get("aggregator_url") else color
        icon_name  = "star"   if r.get("aggregator_url") else "info-sign"

        popup_html = _popup_html(r, idx)
        marker = folium.Marker(
            location=[lat, lon],
            popup=folium.Popup(popup_html, max_width=300),
            tooltip=r.get("name", ""),
            icon=folium.Icon(color=icon_color, icon=icon_name, prefix="glyphicon"),
        )
        _, cluster = q_clusters.get(q, (None, None))
        if cluster is not None:
            marker.add_to(cluster)

    if heat_points:
        HeatMap(heat_points, radius=18, blur=15, min_opacity=0.35).add_to(heat_fg)

    legend_items = "".join(
        f'<div style="margin:3px 0">'
        f'<span style="display:inline-block;width:12px;height:12px;border-radius:50%;'
        f'background:{_folium_color_to_hex(q_color.get(q, "blue"))};margin-right:6px"></span>'
        f'{q} ({sum(1 for r in records if r.get("query") == q)})</div>'
        for q in queries
    )
    legend_html = f"""
    <div style="position:fixed;bottom:30px;left:30px;z-index:9999;
                background:white;padding:12px 16px;border-radius:8px;
                box-shadow:0 2px 10px rgba(0,0,0,.25);font-family:Arial,sans-serif;
                font-size:13px;max-width:200px">
      <b style="font-size:14px">\U0001f4cd \u0417\u0430\u043f\u0440\u043e\u0441\u044b</b><br>
      {legend_items}
      <div style="margin-top:8px;padding-top:8px;border-top:1px solid #eee;font-size:11px;color:#888">
        \u2605 \u2014 \u0447\u0435\u0440\u0435\u0437 taplink/linktree
      </div>
    </div>"""

    info_panel = f"""
    <div style="position:fixed;top:10px;right:10px;z-index:9999;
                background:white;padding:10px 14px;border-radius:8px;
                box-shadow:0 2px 8px rgba(0,0,0,.2);font-family:Arial,sans-serif;font-size:13px">
      <b>\U0001f5fa \u042f\u043d\u0434\u0435\u043a\u0441.\u041a\u0430\u0440\u0442\u044b \u2014 \u043f\u0430\u0440\u0441\u0435\u0440</b><br>
      \u0412\u0441\u0435\u0433\u043e: <b>{len(records)}</b> \u0431\u0438\u0437\u043d\u0435\u0441\u043e\u0432<br>
      \u0413\u043e\u0440\u043e\u0434: {state.CITY}<br>
      <span style="color:#888;font-size:11px">\u041e\u0431\u043d\u043e\u0432\u043b\u0435\u043d\u043e: {datetime.now().strftime("%d.%m.%Y %H:%M")}</span>
    </div>"""

    m.get_root().html.add_child(folium.Element(legend_html))
    m.get_root().html.add_child(folium.Element(info_panel))
    folium.LayerControl(collapsed=False).add_to(m)
    m.save(path)
