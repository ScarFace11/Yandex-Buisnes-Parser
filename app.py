import os
import re
import io
import csv
import json
import queue
import threading
import time
import uuid
import multiprocessing

from flask import Flask, render_template, request, Response, send_from_directory, jsonify

import vk_sender

app = Flask(__name__)
OUTPUT_DIR    = "output"
REVIEWED_FILE = os.path.join(OUTPUT_DIR, "_reviewed.json")

# ── Parallel run state ────────────────────────────────────────
# Each search runs in its own process (multiprocessing) for true isolation.
# A bridge thread reads from the multiprocessing.Queue and forwards to a
# regular queue.Queue that the SSE endpoint can consume with timeout.
_runs: dict[str, dict] = {}
_runs_lock = threading.Lock()


def _new_run() -> dict:
    run_id = uuid.uuid4().hex[:8]
    entry = {
        "id":         run_id,
        "process":    None,       # multiprocessing.Process
        "active":     True,
        "queued":     False,
        "log_queue":  queue.Queue(),  # regular Queue for SSE (fed by bridge)
        "stop_event": threading.Event(),
        "params":     {},
        "files":      [],
        "count":      0,
        "cities":     [],
        "started_at": time.time(),
        "_done":      False,      # bridge signals when process is done
    }
    with _runs_lock:
        _runs[run_id] = entry
    return entry


def _bridge_reader(mp_queue, reg_queue, entry):
    """Read from multiprocessing.Queue, forward to regular Queue for SSE.

    Runs in a daemon thread.  Exits when the process ends and the
    multiprocessing queue is drained (or immediately if process crashed).
    """
    proc = entry.get("process")
    while True:
        try:
            msg = mp_queue.get(timeout=1.0)
            if msg is None:
                break  # sentinel from child
            reg_queue.put(msg)
        except queue.Empty:
            # No message for 1s — check if process is still alive
            if proc and not proc.is_alive():
                # Drain any remaining items
                while not mp_queue.empty():
                    try:
                        msg = mp_queue.get_nowait()
                        if msg is not None:
                            reg_queue.put(msg)
                    except queue.Empty:
                        break
                break
        except Exception:
            break
    entry["_done"] = True


def _finish_run(run_id: str):
    """Mark run done, start next queued run if any."""
    next_entry = None
    with _runs_lock:
        if run_id in _runs:
            _runs[run_id]["active"] = False
        # Atomically find and claim the next queued run
        for rid, r in _runs.items():
            if r.get("queued"):
                r["queued"] = False
                r["active"] = True
                next_entry = r
                break
        # Cleanup old finished runs (> 1 hour)
        now = time.time()
        to_remove = [rid for rid, r in _runs.items()
                     if not r["active"] and not r.get("queued")
                     and now - r.get("started_at", 0) > 3600]
        for rid in to_remove:
            del _runs[rid]
    if next_entry:
        _start_process(next_entry)


def _start_process(entry: dict):
    """Start a search in a child process with bridge thread for SSE."""
    params    = entry["params"]
    run_id    = entry["id"]
    mp_queue  = multiprocessing.Queue()

    # Stop file: child checks periodically and exits gracefully
    stop_dir  = os.path.join(OUTPUT_DIR, ".run_stop")
    os.makedirs(stop_dir, exist_ok=True)
    stop_file = os.path.join(stop_dir, f"{run_id}.stop")

    from yandex_maps_parser.runner import run_process

    proc = multiprocessing.Process(
        target=run_process,
        args=(params, mp_queue, stop_file),
        daemon=True,
    )
    entry["process"] = proc
    entry["stop_file"] = stop_file
    proc.start()

    # Bridge thread: mp_queue → regular queue for SSE
    bridge = threading.Thread(
        target=_bridge_reader,
        args=(mp_queue, entry["log_queue"], entry),
        daemon=True,
    )
    bridge.start()


# ── Reviewed helpers ──────────────────────────────────────────

def _load_reviewed() -> dict:
    if os.path.exists(REVIEWED_FILE):
        try:
            with open(REVIEWED_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _save_reviewed(data: dict):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(REVIEWED_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


_ANSI_RE = re.compile(r"\x1b\[[0-9;]*m")


def _strip_ansi(text: str) -> str:
    return _ANSI_RE.sub("", text)


# ── Routes ────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/run", methods=["POST"])
def run_parser():
    params = request.get_json(force=True) or {}

    queries = [q.strip() for q in params.get("queries", []) if q.strip()]
    if not queries:
        return jsonify({"error": "Введите хотя бы один запрос"}), 400
    cities = [c.strip() for c in params.get("cities", []) if c.strip()]
    if not cities:
        city = params.get("city", "").strip()
        if city:
            cities = [city]
    if not cities:
        return jsonify({"error": "Введите хотя бы один город"}), 400
    params["cities"] = cities

    entry = _new_run()
    entry["params"] = params
    entry["cities"] = cities

    active = _active_run_id()
    if active and active != entry["id"]:
        with _runs_lock:
            entry["queued"] = True
            entry["active"] = False
        return jsonify({"ok": True, "queued": True, "run_id": entry["id"],
                        "position": _queue_position(entry["id"])})

    _start_process(entry)
    return jsonify({"ok": True, "queued": False, "run_id": entry["id"]})


def _queue_position(run_id: str) -> int:
    pos = 1
    with _runs_lock:
        for rid, r in _runs.items():
            if r.get("queued") and rid != run_id:
                pos += 1
    return pos


@app.route("/stop", methods=["POST"])
def stop_parser():
    run_id = request.args.get("run_id", "")
    with _runs_lock:
        targets = []
        if run_id and run_id in _runs:
            targets.append(_runs[run_id])
        else:
            for r in _runs.values():
                if r["active"] and not r.get("queued"):
                    targets.append(r)
        for entry in targets:
            # Signal via stop file
            sf = entry.get("stop_file")
            if sf:
                try:
                    with open(sf, "w") as f:
                        f.write("stop")
                except Exception:
                    pass
            # Also try to terminate the process
            proc = entry.get("process")
            if proc and proc.is_alive():
                proc.terminate()
    return jsonify({"ok": True})


@app.route("/logs")
def logs():
    run_id = request.args.get("run_id", "")
    with _runs_lock:
        if run_id and run_id in _runs:
            entry = _runs[run_id]
        else:
            entry = None
            for r in sorted(_runs.values(), key=lambda x: x["started_at"], reverse=True):
                if r["active"] or r.get("queued"):
                    entry = r
                    break
        if not entry:
            return Response(
                iter(['data: {"type":"ping"}\n\n']),
                mimetype="text/event-stream",
                headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
            )

    lq = entry["log_queue"]

    def generate():
        while True:
            try:
                msg = lq.get(timeout=25)
                yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
                if msg.get("type") == "done":
                    break
            except queue.Empty:
                if entry.get("_done"):
                    # Bridge finished and queue is empty
                    yield 'data: {"type":"done","files":[],"count":0,"stopped":true,"formats":[]}\n\n'
                    break
                yield 'data: {"type":"ping"}\n\n'

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/status")
def status():
    active = _active_run_id()
    queued = sum(1 for r in _runs.values() if r.get("queued"))
    return jsonify({"active": active is not None, "active_run": active, "queued": queued})


@app.route("/runs")
def list_runs():
    with _runs_lock:
        result = []
        for r in _runs.values():
            result.append({
                "id":       r["id"],
                "active":   r["active"],
                "queued":   r.get("queued", False),
                "count":    r.get("count", 0),
                "cities":   r.get("cities", []),
                "files":    r.get("files", []),
            })
        return jsonify({"runs": result})


def _active_run_id() -> str | None:
    with _runs_lock:
        for rid, r in _runs.items():
            if r["active"] and not r.get("queued"):
                return rid
    return None


@app.route("/results/<path:filename>")
def results(filename):
    allowed = os.path.realpath(OUTPUT_DIR)
    filepath = os.path.realpath(os.path.join(allowed, filename))
    if not filepath.startswith(allowed + os.sep) or not os.path.isfile(filepath):
        return jsonify({"error": "not found"}), 404
    try:
        with open(filepath, encoding="utf-8") as f:
            return jsonify(json.load(f))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.route("/download/<path:filename>")
def download(filename):
    return send_from_directory(OUTPUT_DIR, filename, as_attachment=True)


# ── Test API key ──────────────────────────────────────────────

@app.route("/test-api-key", methods=["POST"])
def test_api_key():
    import requests as req_lib
    data = request.get_json(force=True) or {}
    api_key = data.get("api_key", "").strip()
    if not api_key:
        try:
            from config import YANDEX_API_KEY
            api_key = YANDEX_API_KEY
        except Exception:
            return jsonify({"ok": False, "error": "Ключ не задан"}), 400

    try:
        r = req_lib.get(
            "https://geocode-maps.yandex.ru/1.x/",
            params={"apikey": api_key, "format": "json", "geocode": "Москва"},
            timeout=10,
            headers={"User-Agent": "Mozilla/5.0"},
        )
        if r.status_code == 200:
            try:
                geo = r.json()
                found = geo.get("response", {}).get("GeoObjectCollection", {}) \
                           .get("metaDataProperty", {}).get("GeocoderResponseMetaData", {}) \
                           .get("found", "?")
                return jsonify({"ok": True, "status": 200,
                                "message": f"Ключ рабочий. Геокодер нашёл {found} объект(ов) для «Москва»."})
            except Exception:
                return jsonify({"ok": True, "status": 200, "message": "Ключ рабочий (ответ получен)."})
        elif r.status_code == 403:
            return jsonify({"ok": False, "status": 403, "message": "Ключ недействителен или превышен лимит (403)."})
        elif r.status_code == 429:
            return jsonify({"ok": False, "status": 429, "message": "Превышен лимит запросов (429). Попробуйте позже."})
        else:
            return jsonify({"ok": False, "status": r.status_code,
                            "message": f"Неожиданный ответ: HTTP {r.status_code}."})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


# ── Reviewed state ────────────────────────────────────────────

@app.route("/reviewed", methods=["GET"])
def get_reviewed():
    return jsonify(_load_reviewed())

@app.route("/reviewed", methods=["POST"])
def set_reviewed():
    data = request.get_json(force=True) or {}
    url = data.get("url", "").strip()
    state_val = bool(data.get("reviewed", False))
    if not url:
        return jsonify({"error": "url required"}), 400
    reviewed = _load_reviewed()
    if state_val:
        reviewed[url] = True
    else:
        reviewed.pop(url, None)
    _save_reviewed(reviewed)
    return jsonify({"ok": True})


# ── Export filtered rows ──────────────────────────────────────

@app.route("/export-filtered", methods=["POST"])
def export_filtered():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    data = request.get_json(force=True) or {}
    rows = data.get("rows", [])
    fmt  = data.get("format", "csv")

    HEADERS = ["#", "Название", "Категория", "Адрес", "Телефон",
               "Рейтинг", "Отзывов", "VK", "Instagram", "Facebook",
               "Telegram", "YouTube", "TikTok", "OK", "Twitter", "WhatsApp",
               "Другие соцсети", "Агрегатор", "Ссылка", "Запрос"]
    FIELDS = ["_idx", "name", "category", "address", "phone",
              "rating", "reviews", "vk", "instagram", "facebook",
              "telegram", "youtube", "tiktok", "ok", "twitter", "whatsapp",
              "other_socials", "aggregator_url", "yandex_maps_url", "query"]

    if fmt == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(HEADERS)
        for i, row in enumerate(rows, 1):
            writer.writerow([i if f == "_idx" else row.get(f, "") for f in FIELDS])
        csv_bytes = output.getvalue().encode("utf-8-sig")
        return Response(
            csv_bytes,
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=filtered_export.csv"}
        )

    # Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результаты"

    header_fill = PatternFill("solid", fgColor="1A6B3C")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, row in enumerate(rows, 2):
        for col, f in enumerate(FIELDS, 1):
            val = i - 1 if f == "_idx" else row.get(f, "")
            ws.cell(row=i, column=col, value=val)

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 8), 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=filtered_export.xlsx"}
    )


# ── Sender state ──────────────────────────────────────────────

_send_state: dict = {
    "active":     False,
    "log_queue":  queue.Queue(),
    "stop_event": threading.Event(),
}
_send_lock = threading.Lock()

SENDER_CONFIG_FILE = "sender_config.json"


def _load_sender_config() -> dict:
    if os.path.exists(SENDER_CONFIG_FILE):
        try:
            with open(SENDER_CONFIG_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _safe_sender_config(data: dict) -> dict:
    allowed = {
        "message", "delayMin", "delayMax", "limitType", "limitN", "file",
    }
    return {key: data[key] for key in allowed if key in data}


def _save_sender_config(data: dict):
    data = _safe_sender_config(data)
    with open(SENDER_CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _send_thread(params: dict):
    lq         = _send_state["log_queue"]
    stop_event = _send_state["stop_event"]

    def log_fn(level: str, msg: str):
        lq.put({"type": "log", "level": level, "msg": _strip_ansi(msg)})

    try:
        import vk_sender as _vk
        stats = _vk.run_send(params, log_fn, stop_event)
        lq.put({"type": "done", "stats": stats, "stopped": stop_event.is_set()})
    except Exception as exc:
        lq.put({"type": "log",  "level": "warn", "msg": f"Ошибка: {exc}"})
        lq.put({"type": "done", "stats": {}, "stopped": stop_event.is_set()})
    finally:
        with _send_lock:
            _send_state["active"] = False


# ── Sender routes ─────────────────────────────────────────────

@app.route("/send/files")
def send_files():
    from vk_sender.excel_manager import list_excel_files
    files = list_excel_files(OUTPUT_DIR)
    return jsonify({"files": files})


@app.route("/send/config", methods=["GET"])
def send_config_get():
    return jsonify(_safe_sender_config(_load_sender_config()))


@app.route("/send/config", methods=["POST"])
def send_config_post():
    data = request.get_json(force=True) or {}
    _save_sender_config(data)
    return jsonify({"ok": True})


def _safe_excel_filename(filename: str) -> str | None:
    if not filename:
        return None
    basename = os.path.basename(filename)
    resolved = os.path.realpath(os.path.join(OUTPUT_DIR, basename))
    allowed  = os.path.realpath(OUTPUT_DIR)
    if not resolved.startswith(allowed + os.sep) and resolved != allowed:
        return None
    return basename


@app.route("/send/run", methods=["POST"])
def send_run():
    params = request.get_json(force=True) or {}

    safe_file = _safe_excel_filename(params.get("excel_file", ""))
    if not safe_file:
        return jsonify({"error": "Недопустимое имя файла"}), 400

    with _send_lock:
        if _send_state["active"]:
            return jsonify({"error": "Рассылка уже выполняется"}), 409
        _send_state["active"] = True
        _send_state["log_queue"] = queue.Queue()
        _send_state["stop_event"].clear()

    params["excel_file"] = safe_file
    params["output_dir"] = OUTPUT_DIR
    threading.Thread(target=_send_thread, args=(params,), daemon=True).start()
    return jsonify({"ok": True})


@app.route("/send/stop", methods=["POST"])
def send_stop():
    _send_state["stop_event"].set()
    return jsonify({"ok": True})


@app.route("/send/logs")
def send_logs():
    lq = _send_state["log_queue"]

    def generate():
        while True:
            try:
                msg = lq.get(timeout=25)
                yield f"data: {json.dumps(msg, ensure_ascii=False)}\n\n"
                if msg.get("type") == "done":
                    break
            except queue.Empty:
                yield 'data: {"type":"ping"}\n\n'

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/send/status")
def send_status():
    return jsonify({"active": _send_state["active"]})


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)
