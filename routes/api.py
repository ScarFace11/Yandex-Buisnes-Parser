"""Utility routes: reviewed, export, results, download, API key test."""
import os
import io
import csv
import json

from flask import Blueprint, request, Response, send_from_directory, jsonify

OUTPUT_DIR = "output"
REVIEWED_FILE = os.path.join(OUTPUT_DIR, "_reviewed.json")

bp = Blueprint("api", __name__)


def _load_reviewed() -> dict:
    if os.path.exists(REVIEWED_FILE):
        try:
            with open(REVIEWED_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _save_reviewed(data: dict):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    with open(REVIEWED_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


GITHUB_RAW_URL = "https://raw.githubusercontent.com/ScarFace11/Yandex-Buisnes-Parser/main/static/version.json"


@bp.route("/")
def index():
    return __import__("flask").render_template("index.html")


@bp.route("/check-version")
def check_version():
    """Check if a newer version is available on GitHub."""
    import requests as req_lib
    from config import APP_VERSION
    try:
        r = req_lib.get(GITHUB_RAW_URL, timeout=8,
                        headers={"User-Agent": "YandexParser/2.0"})
        if r.status_code == 200:
            remote = r.json()
            remote_ver = remote.get("version", "0.0.0")
            current = APP_VERSION
            # Simple version compare: "2.2.0" > "2.1.0"
            def _ver_tuple(v):
                return tuple(int(x) for x in v.split(".") if x.isdigit())
            newer = _ver_tuple(remote_ver) > _ver_tuple(current)
            return jsonify({
                "current": current,
                "remote": remote_ver,
                "newer": newer,
                "changelog": remote.get("changelog", ""),
                "download_url": "https://github.com/ScarFace11/Yandex-Buisnes-Parser/archive/refs/heads/main.zip",
            })
        return jsonify({"current": APP_VERSION, "newer": False, "error": f"HTTP {r.status_code}"})
    except Exception as exc:
        return jsonify({"current": APP_VERSION, "newer": False, "error": str(exc)})


@bp.route("/reviewed", methods=["GET"])
def get_reviewed():
    return jsonify(_load_reviewed())


@bp.route("/reviewed", methods=["POST"])
def set_reviewed():
    data = request.get_json(force=True) or {}
    url = data.get("url", "").strip()
    state_val = bool(data.get("reviewed", False))
    if not url:
        return jsonify({"error": "url required"}), 400
    reviewed = _load_reviewed()
    if state_val:
        reviewed[url] = True
    else:
        reviewed.pop(url, None)
    _save_reviewed(reviewed)
    return jsonify({"ok": True})


@bp.route("/results/<path:filename>")
def results(filename):
    allowed = os.path.realpath(OUTPUT_DIR)
    filepath = os.path.realpath(os.path.join(allowed, filename))
    if not filepath.startswith(allowed + os.sep) or not os.path.isfile(filepath):
        return jsonify({"error": "not found"}), 404
    try:
        with open(filepath, encoding="utf-8") as f:
            return jsonify(json.load(f))
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@bp.route("/download/<path:filename>")
def download(filename):
    return send_from_directory(OUTPUT_DIR, filename, as_attachment=True)


@bp.route("/save-api-key", methods=["POST"])
def save_api_key():
    """Save YANDEX_API_KEY to .env file and reload it."""
    import re as _re
    data = request.get_json(force=True) or {}
    api_key = data.get("api_key", "").strip()
    if not api_key:
        return jsonify({"ok": False, "error": "Введите ключ API"}), 400

    # Determine .env path — prefer project root
    from pathlib import Path
    project_root = Path(__file__).resolve().parent.parent
    env_path = project_root / ".env"
    if not env_path.exists():
        # Also check yandex_maps_parser/ subfolder
        alt = project_root / "yandex_maps_parser" / ".env"
        if alt.exists():
            env_path = alt

    try:
        # Read existing .env or create new
        lines = []
        if env_path.exists():
            with open(env_path, "r", encoding="utf-8") as f:
                lines = f.readlines()

        # Find and replace or append YANDEX_API_KEY
        found = False
        for i, line in enumerate(lines):
            stripped = line.strip()
            if stripped.startswith("YANDEX_API_KEY") and "=" in stripped:
                lines[i] = f'YANDEX_API_KEY = "{api_key}"\n'
                found = True
                break
        if not found:
            lines.append(f'\nYANDEX_API_KEY = "{api_key}"\n')

        with open(env_path, "w", encoding="utf-8") as f:
            f.writelines(lines)

        # Reload the key into the running config
        os.environ["YANDEX_API_KEY"] = api_key
        try:
            import config
            config.YANDEX_API_KEY = api_key
        except Exception:
            pass

        return jsonify({"ok": True, "message": f"Ключ сохранён в {env_path.name}", "path": str(env_path)})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@bp.route("/export-filtered", methods=["POST"])
def export_filtered():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    data = request.get_json(force=True) or {}
    rows = data.get("rows", [])
    fmt  = data.get("format", "csv")

    HEADERS = ["#", "Название", "Категория", "Адрес", "Телефон",
               "Рейтинг", "Отзывов", "VK", "Instagram", "Facebook",
               "Telegram", "YouTube", "TikTok", "OK", "Twitter", "WhatsApp",
               "Другие соцсети", "Агрегатор", "Ссылка", "Запрос"]
    FIELDS = ["_idx", "name", "category", "address", "phone",
              "rating", "reviews", "vk", "instagram", "facebook",
              "telegram", "youtube", "tiktok", "ok", "twitter", "whatsapp",
              "other_socials", "aggregator_url", "yandex_maps_url", "query"]

    if fmt == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(HEADERS)
        for i, row in enumerate(rows, 1):
            writer.writerow([i if f == "_idx" else row.get(f, "") for f in FIELDS])
        csv_bytes = output.getvalue().encode("utf-8-sig")
        return Response(
            csv_bytes,
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment; filename=filtered_export.csv"}
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Результаты"

    header_fill = PatternFill("solid", fgColor="1A6B3C")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, row in enumerate(rows, 2):
        for col, f in enumerate(FIELDS, 1):
            val = i - 1 if f == "_idx" else row.get(f, "")
            ws.cell(row=i, column=col, value=val)

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 8), 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=filtered_export.xlsx"}
    )


# ── Logs ────────────────────────────────────────────────────

@bp.route("/logs/list")
def list_log_files():
    """List all log files with metadata."""
    from run_logger import list_logs
    return jsonify({"logs": list_logs()})


@bp.route("/logs/view/<path:filename>")
def view_log(filename):
    """View a log file as plain text."""
    from run_logger import LOGS_DIR
    safe = os.path.basename(filename)
    fpath = os.path.join(LOGS_DIR, safe)
    if not os.path.isfile(fpath):
        return jsonify({"error": "Log not found"}), 404
    with open(fpath, encoding="utf-8") as f:
        content = f.read()
    return Response(content, mimetype="text/plain; charset=utf-8",
                    headers={"Content-Disposition": f"inline; filename={safe}"})


@bp.route("/logs/download/<path:filename>")
def download_log(filename):
    """Download a log file."""
    from run_logger import LOGS_DIR
    safe = os.path.basename(filename)
    if not safe.endswith(".log"):
        return jsonify({"error": "Invalid file"}), 400
    fpath = os.path.join(LOGS_DIR, safe)
    if not os.path.isfile(fpath):
        return jsonify({"error": "Log not found"}), 404
    return send_from_directory(LOGS_DIR, safe, as_attachment=True)
