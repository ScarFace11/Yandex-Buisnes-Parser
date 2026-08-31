"""
Run state management: process lifecycle, queue bridging, run registry.

Each search runs in its own process (multiprocessing) for true isolation.
A bridge thread reads from the multiprocessing.Queue and forwards to a
regular queue.Queue that the SSE endpoint can consume with timeout.
"""
import os
import json
import queue
import re
import threading
import time
import uuid
import multiprocessing

OUTPUT_DIR = "output"

# Max time a process can run before we force-kill it (seconds)
_PROCESS_TIMEOUT = 600  # 10 minutes

# Delay before calling finish_run() after bridge exits.
# Gives the frontend time to receive the done message and close the
# SSE connection before a new queued run starts writing to the same queue.
_FINISH_DELAY = 3  # seconds

# Compiled ANSI escape pattern for stripping colour codes from log lines
_ANSI_RE = re.compile(r"\x1b\[[0-9;]*m")


class RunManager:
    """Manages parallel search runs with process isolation."""

    def __init__(self):
        self._runs: dict[str, dict] = {}
        self._lock = threading.Lock()

    def new_run(self) -> dict:
        run_id = uuid.uuid4().hex[:8]
        entry = {
            "id":         run_id,
            "process":    None,
            "active":     True,
            "queued":     False,
            "log_queue":  queue.Queue(),
            "stop_event": threading.Event(),
            "skip_event": threading.Event(),
            "skip_file":  None,
            "params":     {},
            "files":      [],
            "count":      0,
            "cities":     [],
            "started_at": time.time(),
            "_done":      False,
            "_run_mode":  "thread",   # "process" or "thread"
        }
        with self._lock:
            self._runs[run_id] = entry
        return entry

    def get(self, run_id: str) -> dict | None:
        with self._lock:
            return self._runs.get(run_id)

    def active_run_id(self) -> str | None:
        with self._lock:
            for rid, r in self._runs.items():
                if r["active"] and not r.get("queued"):
                    return rid
        return None

    def queue_position(self, run_id: str) -> int:
        pos = 1
        with self._lock:
            for rid, r in self._runs.items():
                if r.get("queued") and rid != run_id:
                    pos += 1
        return pos

    def finish_run(self, run_id: str):
        """Mark run done, start next queued run if any."""
        next_entry = None
        with self._lock:
            if run_id in self._runs:
                self._runs[run_id]["active"] = False
            # Atomically find and claim the next queued run
            for rid, r in self._runs.items():
                if r.get("queued"):
                    r["queued"] = False
                    r["active"] = True
                    next_entry = r
                    break
            # Cleanup old finished runs (> 1 hour)
            now = time.time()
            to_remove = [rid for rid, r in self._runs.items()
                         if not r["active"] and not r.get("queued")
                         and now - r.get("started_at", 0) > 3600]
            for rid in to_remove:
                del self._runs[rid]
        if next_entry:
            self.start_process(next_entry)

    def start_process(self, entry: dict):
        """Start a search in a child process (or thread fallback) with bridge."""
        params   = entry["params"]
        run_id   = entry["id"]
        mp_queue = multiprocessing.Queue()

        stop_dir  = os.path.join(OUTPUT_DIR, ".run_stop")
        os.makedirs(stop_dir, exist_ok=True)
        stop_file = os.path.join(stop_dir, f"{run_id}.stop")
        entry["stop_file"] = stop_file
        skip_file = os.path.join(stop_dir, f"{run_id}.skip")
        entry["skip_file"] = skip_file

        # Try multiprocessing first
        used_process = False
        try:
            from yandex_maps_parser.runner import run_process
            proc = multiprocessing.Process(
                target=run_process,
                args=(params, mp_queue, stop_file, skip_file),
                daemon=True,
            )
            proc.start()
            entry["process"] = proc
            entry["_run_mode"] = "process"
            used_process = True
        except Exception:
            # Fallback: run in a thread (no state isolation, but functional)
            entry["process"] = None   # CRITICAL: clear so bridge uses sentinel
            entry["_run_mode"] = "thread"
            self._start_thread_fallback(entry, mp_queue, stop_file)

        # Bridge thread: mp_queue → regular queue for SSE
        bridge = threading.Thread(
            target=_bridge_reader,
            args=(mp_queue, entry["log_queue"], entry),
            daemon=True,
        )
        bridge.start()

        # Watchdog: force-kill process if it runs too long
        if used_process:
            watchdog = threading.Thread(
                target=_watchdog,
                args=(entry,),
                daemon=True,
            )
            watchdog.start()

    def _start_thread_fallback(self, entry, mp_queue, stop_file):
        """Fallback: run search in a thread when multiprocessing fails."""
        import yandex_maps_parser as _parser

        params = entry["params"]

        def _log_to_queue(level: str, msg: str):
            """Log callback that puts messages into the multiprocessing queue."""
            try:
                clean = _ANSI_RE.sub("", msg)
                if level == "result":
                    # msg is already a JSON string from _emit_result;
                    # parse it to get the dict for the frontend
                    mp_queue.put({"type": "result", "data": json.loads(clean)})
                else:
                    mp_queue.put({"type": "log", "level": level, "msg": clean})
            except Exception:
                pass

        def _thread_run():
            try:
                stop_event = entry["stop_event"]
                skip_event = entry["skip_event"]
                files = _parser.run_web(params, _log_to_queue, stop_event, skip_event)
                # If multiple cities, merge all city JSON files into a combined one
                json_files = [f for f in files if f.endswith(".json") and not f.startswith("_")]
                if len(json_files) > 1:
                    combined = []
                    for jf in json_files:
                        try:
                            fpath = os.path.join(OUTPUT_DIR, jf)
                            with open(fpath, encoding="utf-8") as fh:
                                combined.extend(json.load(fh))
                        except Exception:
                            pass
                    if combined:
                        combined_name = "_combined_results.json"
                        combined_path = os.path.join(OUTPUT_DIR, combined_name)
                        with open(combined_path, "w", encoding="utf-8") as fh:
                            json.dump(combined, fh, ensure_ascii=False, indent=2)
                        files.insert(0, combined_name)
                # Always ensure a results JSON exists for the frontend to load.
                # Check if any user-facing JSON already exists (not internal _ files).
                user_json = [f for f in files if f.endswith(".json") and not f.startswith("_")]
                if not user_json:
                    # No user JSON — create an internal one for the frontend.
                    all_records = []
                    # Try reading from xlsx if available
                    xlsx_files = [f for f in files if f.endswith(".xlsx")]
                    if xlsx_files:
                        try:
                            import openpyxl
                            xlsx_path = os.path.join(OUTPUT_DIR, xlsx_files[0])
                            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
                            ws = wb.active
                            headers = [str(c.value or "") for c in next(ws.iter_rows(max_row=1))]
                            for row in ws.iter_rows(min_row=2, values_only=True):
                                rec = {h: (str(v) if v is not None else "") for h, v in zip(headers, row)}
                                all_records.append(rec)
                            wb.close()
                        except Exception:
                            pass
                    # Fallback: check if _combined_results.json exists
                    if not all_records:
                        combined_path = os.path.join(OUTPUT_DIR, "_combined_results.json")
                        if os.path.exists(combined_path):
                            try:
                                with open(combined_path, encoding="utf-8") as fh:
                                    all_records = json.load(fh)
                            except Exception:
                                pass
                    if all_records:
                        tmp_name = "_results_for_frontend.json"
                        tmp_path = os.path.join(OUTPUT_DIR, tmp_name)
                        with open(tmp_path, "w", encoding="utf-8") as fh:
                            json.dump(all_records, fh, ensure_ascii=False, indent=2)
                        files.insert(0, tmp_name)
                count = 0
                for f in files:
                    if f.endswith(".json"):
                        try:
                            with open(os.path.join(OUTPUT_DIR, f), encoding="utf-8") as jf:
                                count = len(json.load(jf))
                            break
                        except Exception:
                            pass
                fmts = []
                if params.get("output_csv"):   fmts.append("csv")
                if params.get("output_json"):  fmts.append("json")
                if params.get("output_excel"): fmts.append("xlsx")
                if params.get("output_map"):   fmts.append("map")
                # CRITICAL: Send done to mp_queue so the bridge forwards it
                # to the SSE endpoint BEFORE the sentinel. The bridge will
                # exit after seeing the sentinel, and the SSE endpoint will
                # deliver the done message to the frontend.
                # This is the ONLY place done is sent for normal completion.
                skipped = getattr(_parser.state, '_SKIPPED_CITIES', [])
                mp_queue.put({"type": "done", "files": files, "count": count,
                              "stopped": stop_event.is_set(), "formats": fmts,
                              "skipped_cities": skipped})
            except Exception as exc:
                import traceback
                tb = traceback.format_exc()
                try:
                    mp_queue.put({"type": "log", "level": "warn",
                                  "msg": f"Ошибка: {exc}\n{tb}"})
                    mp_queue.put({"type": "done", "files": [], "count": 0,
                                  "stopped": False, "formats": []})
                except Exception:
                    pass
            finally:
                try:
                    mp_queue.put(None)
                except Exception:
                    pass
                if stop_file:
                    try:
                        os.remove(stop_file)
                    except OSError:
                        pass
                skip_f = entry.get("skip_file")
                if skip_f:
                    try:
                        os.remove(skip_f)
                    except OSError:
                        pass

        t = threading.Thread(target=_thread_run, daemon=True)
        entry["_thread"] = t
        t.start()

    def stop_run(self, run_id: str = ""):
        """Stop a run by ID or the active run."""
        with self._lock:
            targets = []
            if run_id and run_id in self._runs:
                targets.append(self._runs[run_id])
            else:
                for r in self._runs.values():
                    if r["active"] and not r.get("queued"):
                        targets.append(r)
            for entry in targets:
                # Signal thread fallback to stop
                stop_ev = entry.get("stop_event")
                if stop_ev:
                    stop_ev.set()
                # Create stop file for multiprocessing mode
                sf = entry.get("stop_file")
                if sf:
                    try:
                        with open(sf, "w") as f:
                            f.write("stop")
                    except Exception:
                        pass
                # Force-kill child process if alive
                proc = entry.get("process")
                if proc and proc.is_alive():
                    try:
                        proc.terminate()
                    except Exception:
                        pass

    def skip_city(self, run_id: str = ""):
        """Skip the current city in the active run."""
        with self._lock:
            targets = []
            if run_id and run_id in self._runs:
                targets.append(self._runs[run_id])
            else:
                for r in self._runs.values():
                    if r["active"] and not r.get("queued"):
                        targets.append(r)
            for entry in targets:
                # Signal thread fallback to skip city
                skip_ev = entry.get("skip_event")
                if skip_ev:
                    skip_ev.set()
                # Create skip file for multiprocessing mode
                sf = entry.get("skip_file")
                if sf:
                    try:
                        with open(sf, "w") as f:
                            f.write("skip")
                    except Exception:
                        pass

    def skip_city_info(self, run_id: str = "") -> dict:
        """Get info about the current city (records found so far)."""
        import yandex_maps_parser.state as _state
        with _state._city_records_lock:
            records = _state._CITY_RECORDS_FOUND
        return {
            "city": _state.CITY,
            "records_found": records,
        }

    def list_runs(self) -> list[dict]:
        with self._lock:
            return [{
                "id":     r["id"],
                "active": r["active"],
                "queued": r.get("queued", False),
                "count":  r.get("count", 0),
                "cities": r.get("cities", []),
                "files":  r.get("files", []),
            } for r in self._runs.values()]

    def status(self) -> dict:
        active = self.active_run_id()
        queued = sum(1 for r in self._runs.values() if r.get("queued"))
        return {"active": active is not None, "active_run": active, "queued": queued}


# ── Singleton ─────────────────────────────────────────────────
run_manager = RunManager()


# ── Bridge thread helper ──────────────────────────────────────

def _bridge_reader(mp_queue, reg_queue, entry):
    """Read from multiprocessing.Queue, forward to regular Queue for SSE.

    Exits when:
    - Child sends None sentinel (normal completion)
    - Child process/thread dies and queue is drained (crash)
    - Exception occurs

    CRITICAL DESIGN:
    - The done message is ONLY sent when the producer is confirmed dead
      (crash case). For normal completion, the producer itself sends
      the done message before the sentinel.
    - finish_run() is delayed by _FINISH_DELAY seconds after the bridge
      exits. This gives the frontend time to receive the done message
      and close the SSE connection before a new queued run starts.
    """
    mode = entry.get("_run_mode", "thread")
    proc = entry.get("process")
    thread = entry.get("_thread")
    empty_count = 0  # consecutive empty polls

    while True:
        try:
            msg = mp_queue.get(timeout=2.0)
            if msg is None:
                break  # sentinel — child is done
            reg_queue.put(msg)
            empty_count = 0
        except queue.Empty:
            empty_count += 1
            # Check if the producer is still alive
            alive = False
            if mode == "process" and proc:
                alive = proc.is_alive()
            elif mode == "thread" and thread:
                alive = thread.is_alive()
            else:
                # Unknown mode — assume alive until sentinel
                alive = True

            if not alive:
                # Producer died — drain any remaining messages and exit
                while not mp_queue.empty():
                    try:
                        msg = mp_queue.get_nowait()
                        if msg is not None:
                            reg_queue.put(msg)
                    except queue.Empty:
                        break
                break

            # Safety: if idle for >60s with no sentinel, something is wrong
            if empty_count >= 30:
                # Drain remaining messages
                while not mp_queue.empty():
                    try:
                        msg = mp_queue.get_nowait()
                        if msg is not None:
                            reg_queue.put(msg)
                    except queue.Empty:
                        break
                break

        except Exception:
            break

    # CRASH SAFETY: If the producer died without sending a done message,
    # send one now so the frontend gets a clean completion signal.
    if not entry.get("_done"):
        try:
            reg_queue.put({"type": "done", "files": entry.get("files", []),
                           "count": entry.get("count", 0), "stopped": False, "formats": [],
                           "skipped_cities": []})
        except Exception:
            pass

    entry["_done"] = True

    # CRITICAL: Delay finish_run() to give the frontend time to receive
    # the done message and close the SSE connection. Without this delay,
    # a queued run could start and write to the same SSE queue before
    # the frontend disconnects, causing results to be lost.
    _rid = entry["id"]
    threading.Timer(_FINISH_DELAY, lambda: run_manager.finish_run(_rid)).start()


def _watchdog(entry):
    """Force-kill a process if it exceeds _PROCESS_TIMEOUT."""
    start = entry["started_at"]
    while True:
        time.sleep(5)
        elapsed = time.time() - start
        if elapsed > _PROCESS_TIMEOUT:
            proc = entry.get("process")
            if proc and proc.is_alive():
                try:
                    proc.terminate()
                except Exception:
                    pass
            break
        # Stop watching if run already finished
        if entry.get("_done"):
            break
