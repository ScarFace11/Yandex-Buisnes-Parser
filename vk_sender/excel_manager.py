"""
Работа с Excel-файлами парсера:
  - Поиск файлов в output/
  - Чтение получателей (строк с заполненной соцсетью без пометки «Отправлено»)
  - Запись статуса отправки
"""
import os
import logging
from typing import Generator, Optional

import openpyxl

logger = logging.getLogger(__name__)

# Варианты написания заголовка колонки VK в Excel-файлах парсера
_VK_HEADER_VARIANTS = {"вконтакте", "vk", "вк", "vkontakte"}

# Заголовок колонки статуса отправки
SENT_HEADER = "Отправлено"


def list_excel_files(output_dir: str = "output") -> list[str]:
    """Вернуть список .xlsx файлов в output_dir (только имена, не пути)."""
    if not os.path.isdir(output_dir):
        return []
    return sorted(
        f for f in os.listdir(output_dir)
        if f.endswith(".xlsx") and not f.startswith("_") and not f.startswith("~$")
    )


def _find_col(ws, header: str) -> Optional[int]:
    """Найти номер столбца (1-based) по заголовку в первой строке."""
    norm = header.strip().lower()
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().lower() == norm:
            return cell.column
    return None


def _find_vk_col(ws) -> Optional[int]:
    """Найти столбец ВКонтакте по нескольким вариантам написания."""
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().lower() in _VK_HEADER_VARIANTS:
            return cell.column
    return None


def _get_or_create_sent_col(ws) -> int:
    """Вернуть номер колонки «Отправлено», создав её если нет."""
    col = _find_col(ws, SENT_HEADER)
    if col:
        return col
    # Добавить в конец
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col, value=SENT_HEADER)
    return new_col


def _find_name_col(ws) -> Optional[int]:
    """Найти колонку «Название» (name)."""
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().lower() in ("название", "name"):
            return cell.column
    return None


def iter_recipients(
    excel_path: str,
    social: str = "vk",
) -> Generator[dict, None, None]:
    """
    Итерировать по всем кандидатам из Excel-файла.

    Для каждой строки, где:
      - есть ссылка на нужную соцсеть
      - в колонке «Отправлено» нет «+»

    Yields dict с ключами:
      row_num (int), name (str), vk_url (str), sent (str)

    Ограничение по числу записей (limit) применяется в runner.py
    ПОСЛЕ проверок сайта и истории, чтобы N означало N фактически
    отправленных сообщений, а не N строк Excel.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    vk_col   = _find_vk_col(ws) if social == "vk" else _find_col(ws, social)
    sent_col = _find_col(ws, SENT_HEADER)
    name_col = _find_name_col(ws)

    if not vk_col:
        logger.warning("Колонка ВКонтакте не найдена в %s", excel_path)
        wb.close()
        return

    for row_num in range(2, ws.max_row + 1):
        vk_url = ws.cell(row=row_num, column=vk_col).value
        if not vk_url or not str(vk_url).strip():
            continue

        sent_val = ""
        if sent_col:
            sent_val = str(ws.cell(row=row_num, column=sent_col).value or "").strip()

        if sent_val == "+":
            continue  # уже отправлено

        name = ""
        if name_col:
            name = str(ws.cell(row=row_num, column=name_col).value or "").strip()

        yield {
            "row_num":  row_num,
            "name":     name or f"Бизнес #{row_num - 1}",
            "vk_url":   str(vk_url).strip(),
            "sent":     sent_val,
        }

    wb.close()


def mark_sent(
    excel_path: str,
    row_num: int,
    status: str = "+",
) -> None:
    """
    Проставить статус в колонке «Отправлено» для строки row_num.
    Создаёт колонку если её нет.
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    sent_col = _get_or_create_sent_col(ws)
    ws.cell(row=row_num, column=sent_col, value=status)
    wb.save(excel_path)
    wb.close()
